
import openpyxl
import os

def write_to_excel(filename, video_number, num_images):
    
    if not os.path.isfile(filename):
    
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1,column=1,value='Video Number')
        worksheet.cell(row=1,column=2, value='Number of Images')
    else:  
        # Load the workbook
        workbook = openpyxl.load_workbook(filename)
    
        # Select the active worksheet
        worksheet = workbook.active

    last_row = worksheet.max_row 
    
    # Write the data to the worksheet
    worksheet.cell(row=last_row+1, column=1, value=video_number)
    worksheet.cell(row=last_row+1, column=2, value=num_images)

    # Save the workbook 
    workbook.save(filename)

def find_min_images(filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)

    # Select the active worksheet 
    worksheet = workbook.active

    # Initialize the minimum number of images
    min_images = float('inf')
    
    print("Rows of excel file: \n")

    # Iterate over the rows in the worksheet
    for row in worksheet.iter_rows(min_row=2, max_col=2):
        # Get the number of images for the current video file
        print(row)  
        num_images = row[1].value

        # Update the minimum number of images if necessary
        if num_images < min_images:
            min_images = num_images

    # Return the minimum number of images
    return min_images






