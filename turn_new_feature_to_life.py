# -*- coding: utf-8 -*-
"""
Created on Mon Aug 14 22:37:28 2023

@author: marco
"""

import PySimpleGUI as sg
import importlib
import inspect
import numpy as np  # Import numpy for data manipulation
import openpyxl 

# Function to get the number of independent variables of a function
def get_independent_variables(func):
    try:
        sig = inspect.signature(func)
        return len([param for param in sig.parameters.values() if param.default == param.empty])
    except Exception as e:
        return 0

# Load functions from the "extra.py" module
import extra  # Make sure you have a file named "extra.py" with functions defined

# Get a list of available functions in "extra.py" that contain "_feature" in their names
available_functions = [name for name, obj in inspect.getmembers(extra) if inspect.isfunction(obj) and "_feature" in name]

def find_string_in_sheet(sheet, target_string, eq_signal):
    print("T1")
    for row in sheet.iter_rows():         
        for cell in row:
            print("T2")
            cell_value = cell.value
            print("T3")
            print("Cell Value: " + str(type(cell_value)))
            print(cell_value)
            print("target_string: " + str(target_string))
            
            if cell_value and isinstance(cell_value, str) and target_string in cell_value and eq_signal in cell_value:
                print("T4")
                return cell.coordinate, cell_value
     
def is_first_cell_expressions_list(file_path, word_f, word_eq):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        first_cell_value = sheet.cell(row=1, column=1).value
        if first_cell_value == "Expressions list":
            print("Yes") 
            sheet = workbook.active      
            print("After 1")
            cell_code_number, cell_value = find_string_in_sheet(sheet, word_f, word_eq)
            print("cell_value in that function: ")
            print(cell_value)
            
            # import sys
            # sys.exit()
            
            print("After 2")
            if cell_code_number:
                sg.popup(f"Found feature definition at cell {cell_code_number} with expression: {cell_value}")
            else:
                sg.popup_error("The string for that feature formula was not found in this sheet.")
            
            print("Cell Value: "+ str(cell_value))
            
            # import sys
            # sys.exit()
            
            return (cell_code_number, cell_value)
        else:
            return None
    except Exception as e:
        return None
    
def compute_function_values(function, feature_values):
    print("Computing") 
    import importlib
    import inspect
    print("Computing 2")
    # Import the module dynamically
    module = importlib.import_module("extra")
    
    print("Here A")
    
    # Get the function by its name
    my_function = getattr(module, function)
    
    indSep = 0
    
    for indP, p in enumerate(function):
        if p == '_' :
            indSep = indP
            break
    word_f = function[:indSep]
    
    word_eq = '='
    
    ## GUI to select the excel file with the equations
    ## for the excel file, search for the name of the feature, and extract the expression from there
    
    layout = [
        [sg.Text("Select an Excel file with features expressions:")],
        [sg.InputText(key="-FILE-"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
        [sg.Button("Ok"), sg.Button("Exit")]
    ]

    window = sg.Window("Excel File Checker", layout)
    
    cell_value = ""

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == "Exit":
            break

        if event == "Ok":
            file_path = values["-FILE-"]

            tupInfo = is_first_cell_expressions_list(file_path, word_f, word_eq)
            
            if tupInfo != None:
                sg.popup("First cell contains 'Expressions list'.")
                
                
                    
                cell_code_number, cell_value = tupInfo
                 
                break
            else:
                sg.popup_error("The first cell does not contain 'Expressions list'. Please select another file.")

    window.close()
    
    print("Cell value: " + str(cell_value))

    
    # import sys
    # sys.exit()
     
    ## Check which cell, in the excel file, contains feature_name. For that cell, returns the expressions inside it
    
    
    ####################################
    print("feature_values: ")
    print(feature_values)
    
    l = 0
    
    for i, f in enumerate(feature_values):
        l = len(f)
    
    featAll = []
    
    for x in range(0,l):
        featHere = []
        for i, f in enumerate(feature_values):
            featHere.append(f[x])
            
        featAll.append(featHere)       
    
    print("Here B")
    
    func_args = inspect.getfullargspec(my_function).args
    arg_count = len(func_args)
    arg_count -= 1 
    
    print("Here C")
    
    print(featAll)
    
    results = [] 
    for indS, feature_set in enumerate(featAll):
        print("inds: " + str(indS))
        print("feature_set: ")
        print(tuple(feature_set))
        
        args_dict = dict(zip(func_args, feature_set))
        
        print("args_dict: ")
        print(args_dict) 
        
     #   args_dict['expression'] = cell_value
        
 #       print("args_dict after: ")
 #       print(args_dict)
 
 
        print("Data:")
        for f in feature_set:
            print(f)
        print(cell_value)
        
        indeq = 0
        
        # for indc, c in enumerate(cell_value):
        #     if c == '=':
        #       indeq = indc
        #       break
    
        a = False
    
        if '=' in cell_value:            
            cell_value_sep = cell_value.split('=')
            a = True
        else:
            cell_value_sep = cell_value
        
        if a:        
            cell_value = cell_value_sep[1]
        
        print("Cell value: " + str(cell_value))
        
        # import sys
        # sys.exit()
 
        if len(feature_set) == 1:
            print("\n\n Ok \n\n")
            result = my_function(feature_set[0], cell_value)
        elif len(feature_set) == 2:
            print("\n\n Ok \n\n")
            result = my_function(feature_set[0], feature_set[1], cell_value)
        elif len(feature_set) == 3:
            print("\n\n Ok \n\n")
            result = my_function(feature_set[0], feature_set[1], feature_set[2], cell_value)
        elif len(feature_set) == 4:
            print("\n\n Ok \n\n")
            print("(" + str(feature_set[0]) + " , " + str(feature_set[1]) + " , " + str(feature_set[2]) + " , " + str(feature_set[3]) + " , " + cell_value + ")")
      #      result = my_function(1.788659194757832, 1, 0.5168, 0.999999998263502, 'O = a*b + c*d')
            result = my_function(feature_set[0], feature_set[1], feature_set[2], feature_set[3], cell_value)
        elif len(feature_set) == 5:
            print("\n\n Ok \n\n")
            result = my_function(feature_set[0], feature_set[1], feature_set[2], feature_set[3], feature_set[4], cell_value)
        elif len(feature_set) == 6:
            print("\n\n Ok \n\n")
            result = my_function(feature_set[0], feature_set[1], feature_set[2], feature_set[3], feature_set[4], feature_set[5], cell_value)
        elif len(feature_set) == 7:
            print("\n\n Ok \n\n")
            result = my_function(feature_set[0], feature_set[1], feature_set[2], feature_set[3], feature_set[4], feature_set[5], feature_set[6], cell_value)
                
        
        print("Result: ")
        print(result)
        
        # import sys
        # sys.exit()
        
        results.append(result)
    return results 

def feature_life(example_feature_names, feature_values):
    
    global_done = False
    this_closes = False
    
    # Create the PySimpleGUI layout for function selection
    function_selection_layout = [
        [sg.Text("Select an option:")],
        [sg.Radio("Input parameter values", "RADIO1", default=True, key="-INPUT-VALUES-")],
        [sg.Radio("Single values for each parameter", "RADIO1", key="-SINGLE-VALUES-")],
        [sg.Radio("Associate with existing features", "RADIO1", key="-EXISTING-FEATURES-")],
        [sg.Button("Next"), sg.Button("Exit")]
    ]
 
    # Create the main PySimpleGUI window with option selection layout
    window = sg.Window("Function Call Interface", function_selection_layout)

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == "Exit":
            break

        if event == "Next": 
            if values["-INPUT-VALUES-"]:
                
                again_feat = True
                
                while again_feat:
                
                    func_args = []
                    
                    selected_functions = []
    
                    # Create the PySimpleGUI layout for existing feature function selection with checkboxes
                    existing_feature_layout = [
                        [sg.Text("Select existing feature functions:")],
                        *[[sg.Checkbox(function_name, key=f"-CHECKBOX-{i}-")] for i, function_name in enumerate(available_functions)],
                        [sg.Button("Next"), sg.Button("Exit")]
                    ]
    
                    # Create a window for existing feature function selection
                    existing_feature_window = sg.Window("Existing Feature Function Selection", existing_feature_layout)
    
                    while True:
                        feature_event, feature_valuesx = existing_feature_window.read()
    
                        if feature_event == sg.WIN_CLOSED or feature_event == "Exit":
                            again_feat = True
                            break 
    
                        if feature_event == "Next":
                            again_feat = False
                            # Check each checkbox to see if it's selected
                            for i, function_name in enumerate(available_functions):
                                if feature_valuesx[f"-CHECKBOX-{i}-"]:
                                    selected_functions.append(function_name)
    
                            if not selected_functions:
                                sg.popup_error("Please select at least one existing feature function.")
                                continue
     
                            selected_function_name = selected_functions[0]  # For simplicity, select the first function
                            selected_function = getattr(extra, selected_function_name)
                            
                            print("Computing") 
                            import importlib
                            import inspect
                            print("Computing 2")
                            # Import the module dynamically
                            module = importlib.import_module("extra")
                            
                            print("Here A")
                            
                            # Get the function by its name
                            my_function = getattr(module, selected_function_name)
                             
                            print("A")
                            
                            func_args = inspect.getfullargspec(my_function).args
                            
                            if len(func_args) > 0:
                                break
                        
                # function_name = sg.popup_get_text("Enter the function name:")
                # module_name = sg.popup_get_text("Enter the module name (without .py):")

                # params = get_independent_variables(getattr(extra, function_name))
                input_values = []
                
                func_args = func_args[:-1]

                if not func_args:
                    sg.popup_error("Function or module not found.")
                    continue

                for param in func_args:
                    value = sg.popup_get_text(f"Enter value for parameter '{param}':")
                    input_values.append(eval(value.strip()) if value.strip() else None)
                
                print("Input values: ")
                print(input_values)
                
                list_input = []
                
                for ind in range(len(input_values[0])):
                    list_input.append([x[ind] for x in input_values])
                
                # import sys
                # sys.exit() 
                
                # module_name = "extra.py"
                
                # # Call the function with input values
                # module = importlib.import_module(module_name)
                # function = getattr(module, function_name)
                
                list_for_f = []
                
                for i in input_values:
                    list_for_f.append(i)
                
                feature_values.append(list_for_f)
                
                indSep = 0
                
                for indP, p in enumerate(selected_function_name):
                    if p == '_' :
                        indSep = indP
                        break
                word_f = selected_function_name[:indSep]
                
                word_eq = '='
                
                layout = [
                    [sg.Text("Select an Excel file with features expressions:")],
                    [sg.InputText(key="-FILE-"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
                    [sg.Button("Ok"), sg.Button("Exit")]
                ]

                window = sg.Window("Excel File Checker", layout)
                
                cell_value = ""

                while True:
                    event, values = window.read()

                    if event == sg.WIN_CLOSED or event == "Exit":
                        break

                    if event == "Ok":
                        file_path = values["-FILE-"]

                        tupInfo = is_first_cell_expressions_list(file_path, word_f, word_eq)
                        
                        if tupInfo != None:
                            sg.popup("First cell contains 'Expressions list'.")
                            
                            
                                
                            cell_code_number, cell_value = tupInfo
                             
                            break
                        else:
                            sg.popup_error("The first cell does not contain 'Expressions list'. Please select another file.")

                window.close()
                
                print("Cell value: " + str(cell_value))
                
                cell_value_sp = cell_value.split('=')
                cell_value = cell_value_sp[1]
                
                print("Cell value: " + str(cell_value))
                    
        ##        results = compute_function_values(selected_function_name, feature_values)
            #    result = function(*input_values)
                
                results = []
            
                for input_values in list_input:
                    
                    print("Input values: ")
                    print(input_values)
                    
                    print(input_values[0])
                    print(input_values[1])
                    print(input_values[2])
                    print(input_values[3])
             
                    if len(input_values) == 1:
                        print("\n\n Ok \n\n")
                        result = my_function(input_values[0], cell_value)
                    elif len(input_values) == 2:
                        print("\n\n Ok \n\n")
                        result = my_function(input_values[0], input_values[1], cell_value)
                    elif len(input_values) == 3:
                        print("\n\n Ok \n\n")
                        result = my_function(input_values[0], input_values[1], input_values[2], cell_value)
                    elif len(input_values) == 4:
                        print("\n\n Ok \n\n")                   
                        result = my_function(input_values[0], input_values[1], input_values[2], input_values[3], cell_value)
                    elif len(input_values) == 5:
                        print("\n\n Ok \n\n")
                        result = my_function(input_values[0], input_values[1], input_values[2], input_values[3], input_values[4], cell_value)
                    elif len(input_values) == 6:
                        print("\n\n Ok \n\n")
                        result = my_function(input_values[0], input_values[1], input_values[2], input_values[3], input_values[4], input_values[5], cell_value)
                    elif len(input_values) == 7:
                        print("\n\n Ok \n\n")
                        result = my_function(input_values[0], input_values[1], input_values[2], input_values[3], input_values[4], input_values[5], input_values[6], cell_value)
                    
                    print("result: " + str(result))
                    
                    results.append(result)
                
                # import sys
                # sys.exit()
                    
                for result in results:
                    print(feature_values)
                    
                    featValNew = []
                    
                    for f in feature_values:
                        f += (result,)
                        featValNew.append(f)
                    
                    feature_values = featValNew
                        
                    print(feature_values)
                
                print("feature_values: ")
                print(feature_values)
                
                feature_values = feature_values[:-1]
                
                # import sys
                # sys.exit()
                
                sg.popup(f"{len(results)} constant features added to features dataset !!!")          
                
                results = feature_values

            elif values["-SINGLE-VALUES-"]:
                
                again_feat = True
                
                while again_feat:
                
                    func_args = []
                    
                    selected_functions = []
    
                    # Create the PySimpleGUI layout for existing feature function selection with checkboxes
                    existing_feature_layout = [
                        [sg.Text("Select existing feature functions:")],
                        *[[sg.Checkbox(function_name, key=f"-CHECKBOX-{i}-")] for i, function_name in enumerate(available_functions)],
                        [sg.Button("Next"), sg.Button("Exit")]
                    ]
    
                    # Create a window for existing feature function selection
                    existing_feature_window = sg.Window("Existing Feature Function Selection", existing_feature_layout)
    
                    while True:
                        feature_event, feature_valuesx = existing_feature_window.read()
    
                        if feature_event == sg.WIN_CLOSED or feature_event == "Exit":
                            again_feat = True
                            break 
    
                        if feature_event == "Next":
                            again_feat = False
                            # Check each checkbox to see if it's selected
                            for i, function_name in enumerate(available_functions):
                                if feature_valuesx[f"-CHECKBOX-{i}-"]:
                                    selected_functions.append(function_name)
    
                            if not selected_functions:
                                sg.popup_error("Please select at least one existing feature function.")
                                continue
    
                            selected_function_name = selected_functions[0]  # For simplicity, select the first function
                            selected_function = getattr(extra, selected_function_name)
                            
                            print("Computing") 
                            import importlib
                            import inspect
                            print("Computing 2")
                            # Import the module dynamically
                            module = importlib.import_module("extra")
                            
                            print("Here A")
                            
                            # Get the function by its name
                            my_function = getattr(module, selected_function_name)
                             
                            print("A")
                            
                            func_args = inspect.getfullargspec(my_function).args
                            
                            if len(func_args) > 0:
                                break
                        
                # function_name = sg.popup_get_text("Enter the function name:")
                # module_name = sg.popup_get_text("Enter the module name (without .py):")

                # params = get_independent_variables(getattr(extra, function_name))
                input_values = []
                
                func_args = func_args[:-1]

                if not func_args:
                    sg.popup_error("Function or module not found.")
                    continue

                for param in func_args:
                    value = sg.popup_get_text(f"Enter value for parameter '{param}':")
                    input_values.append(eval(value.strip()) if value.strip() else None)
                
                print("Input values: ")
                print(input_values)

                list_for_f = []
                
                for i in input_values:
                    list_for_f.append(i)
                
       #         feature_values.append(list_for_f)
                
                indSep = 0
                
                for indP, p in enumerate(selected_function_name):
                    if p == '_' :
                        indSep = indP
                        break
                word_f = selected_function_name[:indSep]
                
                word_eq = '='
                
                layout = [
                    [sg.Text("Select an Excel file with features expressions:")],
                    [sg.InputText(key="-FILE-"), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
                    [sg.Button("Ok"), sg.Button("Exit")]
                ]

                window = sg.Window("Excel File Checker", layout)
                
                cell_value = ""

                while True:
                    event, values = window.read()

                    if event == sg.WIN_CLOSED or event == "Exit":
                        break

                    if event == "Ok":
                        file_path = values["-FILE-"]

                        tupInfo = is_first_cell_expressions_list(file_path, word_f, word_eq)
                        
                        if tupInfo != None:
                            sg.popup("First cell contains 'Expressions list'.")
                            
                            
                                
                            cell_code_number, cell_value = tupInfo
                             
                            break
                        else:
                            sg.popup_error("The first cell does not contain 'Expressions list'. Please select another file.")

                window.close()
                
                print("Cell value: " + str(cell_value))
                
                cell_value_sp = cell_value.split('=')
                cell_value = cell_value_sp[1]
                
                print("Cell value: " + str(cell_value))
                    
        ##        results = compute_function_values(selected_function_name, feature_values)
            #    result = function(*input_values)
            
                if len(input_values) == 1:
                    print("\n\n Ok \n\n")
                    result = my_function(input_values[0], cell_value)
                elif len(input_values) == 2:
                    print("\n\n Ok \n\n")
                    result = my_function(input_values[0], input_values[1], cell_value)
                elif len(input_values) == 3:
                    print("\n\n Ok \n\n")
                    result = my_function(input_values[0], input_values[1], input_values[2], cell_value)
                elif len(input_values) == 4:
                    print("\n\n Ok \n\n")                   
                    result = my_function(input_values[0], input_values[1], input_values[2], input_values[3], cell_value)
                elif len(input_values) == 5:
                    print("\n\n Ok \n\n")
                    result = my_function(input_values[0], input_values[1], input_values[2], input_values[3], input_values[4], cell_value)
                elif len(input_values) == 6:
                    print("\n\n Ok \n\n")
                    result = my_function(input_values[0], input_values[1], input_values[2], input_values[3], input_values[4], input_values[5], cell_value)
                elif len(input_values) == 7:
                    print("\n\n Ok \n\n")
                    result = my_function(input_values[0], input_values[1], input_values[2], input_values[3], input_values[4], input_values[5], input_values[6], cell_value)
                
                print(feature_values)
                
                featValNew = []
                
                for f in feature_values:
                    f += (result,)
                    featValNew.append(f)
                
                feature_values = featValNew
                    
                print(feature_values)
             ##   print(len(feature_values[0]))
              
                
                # import sys
                # sys.exit()
                 
                results = feature_values
         ##       sg.popup(f"Function result: {result}")
                
                sg.popup(f"1 constant feature added to features dataset !!!")               

            elif values["-EXISTING-FEATURES-"]:
                
                again_feat = True
                
                while again_feat:
                    
                    selected_functions = []
    
                    # Create the PySimpleGUI layout for existing feature function selection with checkboxes
                    existing_feature_layout = [
                        [sg.Text("Select existing feature functions:")],
                        *[[sg.Checkbox(function_name, key=f"-CHECKBOX-{i}-")] for i, function_name in enumerate(available_functions)],
                        [sg.Button("Next"), sg.Button("Exit")]
                    ]
    
                    # Create a window for existing feature function selection
                    existing_feature_window = sg.Window("Existing Feature Function Selection", existing_feature_layout)
    
                    while True:
                        feature_event, feature_valuesx = existing_feature_window.read()
    
                        if feature_event == sg.WIN_CLOSED or feature_event == "Exit":
                            again_feat = True
                            break 
    
                        if feature_event == "Next":
                            again_feat = False
                            # Check each checkbox to see if it's selected
                            for i, function_name in enumerate(available_functions):
                                if feature_valuesx[f"-CHECKBOX-{i}-"]:
                                    selected_functions.append(function_name)
    
                            if not selected_functions:
                                sg.popup_error("Please select at least one existing feature function.")
                                continue
    
                            selected_function_name = selected_functions[0]  # For simplicity, select the first function
                            selected_function = getattr(extra, selected_function_name)
                            num_independent_variables = get_independent_variables(selected_function)
                            
                            print("Computing") 
                            import importlib
                            import inspect
                            print("Computing 2")
                            # Import the module dynamically
                            module = importlib.import_module("extra")
                            
                            print("Here A")
                            
                            # Get the function by its name
                            my_function = getattr(module, selected_function_name)
                             
                            print("A")
                            
                            func_args = inspect.getfullargspec(my_function).args
                            arg_count = len(func_args)
                            arg_count -= 1 
                            
                            print("B")
                            
                            num_independent_variables = arg_count
                            
                            if example_feature_names[0] == 'x':
                                example_feature_names = example_feature_names[1:]
    
                            print("num_independent_variables: " + str(num_independent_variables))
                            
                            # Create the PySimpleGUI layout for variable selection
                            variable_selection_layout = [
                                [sg.Text(f"Select {num_independent_variables} variables for '{selected_function_name}':")],
                                *[[sg.Checkbox(feature_name, key=f"-CHECKBOX-{i}-")] for i, feature_name in enumerate(example_feature_names)],
                                [sg.Button("Next"), sg.Button("Exit")]
                            ]
    
                            # Create a window for variable selection
                            variable_window = sg.Window("Variable Selection", variable_selection_layout)
    
                            while True:
                                var_event, var_values = variable_window.read()
    
                                if var_event == sg.WIN_CLOSED or var_event == "Exit":
                                    break
    
                                if var_event == "Next":
                                    selected_variables = [feature_names for i, feature_names in enumerate(example_feature_names) if var_values[f"-CHECKBOX-{i}-"]]
                                    
                                    if len(selected_variables) != num_independent_variables:
                                        sg.popup_error("The number of selected variables must match the number of independent variables in the function.")
                                    else:
                                        # You can now proceed with the selected variables for the function
                                        sg.popup(f"Selected variables for '{selected_function_name}': {selected_variables}")
                                                                    
                                    break    
    
                            variable_window.close()
                            
                            print("Here C")
    
                            # Now, let's generate some sample feature values (you can replace this with your table data)
                     #       num_samples = 50
                    #        feature_values = np.random.rand(num_samples, len(example_feature_names))
                            
                            print(selected_function)
                            
                            print("Selected Variables: ")
                            print(selected_variables)
                            
                            print("feature_values")
                            print(feature_values)
                            print("example_feature_names")
                            print(example_feature_names)
                        
                  #          import numpy as np
                            
                 #           import sys
                 #           sys.exit()
                 
                            list_for_f = []
                 
                            for s in selected_variables:
                                
                                print("A")
                 
                                w = np.where(np.array(example_feature_names) == s)
                                print("w: " + str(w[0][0]))
                                list_for_w = []
                                
                                for fv in feature_values:
                                    list_for_w.append(fv[w[0][0]])
                                
                                list_for_f.append(list_for_w)
                            
                            print("List for f:")
                            print(list_for_f)
                            
                            # import sys
                            # sys.exit()
    
                            # Compute the function values for the selected function and feature values
                            results = compute_function_values(selected_function_name, list_for_f)
                            print("results: ")
                            print(results)
                            
                            featValNewHere = []
                            
                            for indF, f in enumerate(feature_values):
                                print("f: ")
                                print(f)
                                print(results[indF])
                                f += (results[indF],)
                                print("f: ")
                                print(f)
                                
                                # import sys
                                # sys.exit()
                            
                                featValNewHere.append(f)
                                
                            results = featValNewHere
                            
                            print("results: ")
                            print(results)
                            
                            # import sys
                            # sys.exit()
                            
                       #     sg.popup(f"Computed results for '{selected_function_name}':\n{results}")
                            global_done = True
                            again_feat = False
                            break  # Exit the loop if at least one feature function is selected
    
                    existing_feature_window.close()
                    
                    if global_done:
                        global_done = False
                        this_closes = True
                        break
            if this_closes:
                break

    window.close()
    
    return results




