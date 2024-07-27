# -*- coding: utf-8 -*-
"""
Created on Sun Jul 30 14:38:33 2023

@author: marco
"""


def features_auto(existing_features_list):
    
    print("In 1")

    import PySimpleGUI as sg
    print("Nice 1")
    import sympy as sp
    print("Nice 2")
  #  import extra  # Import the functions from extra.py
    print("Nice 3")
    import os
    print("Nice 4")
    import openpyxl
     
    print("In 2")

    # ... (preprocess_formula, is_valid_formula, create_feature_function functions) ...

    def preprocess_formula(expression):
        return expression.replace("^2", "**2").replace("^", "**")

    def is_valid_formula(expression, variables, variable_values):
        try:
            processed_expression = preprocess_formula(expression)
            var_dict = {var: float(val) for var, val in zip(variables.split(","), variable_values.split(","))}
            sp.sympify(processed_expression, locals=var_dict)
            return True
        except sp.SympifyError:  
            return False
        except ValueError:
            return False

    # def create_feature_function(dependent_var, expression, independent_vars):
    #     # Generate the function body dynamically
    #     func_body = f"def {dependent_var}_feature({', '.join(independent_vars)}):\n"
    #     func_body += "    var_dict = {var: float(val) for var, val in zip(independent_vars, args)}\n"
    #     func_body += "    processed_expression = preprocess_formula(expression)\n"
    #     func_body += "    return float(sp.sympify(processed_expression, locals=var_dict))\n"

    #     return func_body
    
    
    def write_to_excel(filename, cell_value):
        # Load the existing Excel file or create a new one if it doesn't exist
        try:
            workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Select the first (or active) worksheet
        sheet = workbook.active

        # Check if the first cell is empty, and if so, write "Expressions list" to it
        if sheet.cell(row=1, column=1).value is None:
            sheet.cell(row=1, column=1, value="Expressions list")

        # Find the next available row in column A
        next_row = sheet.max_row + 1 

        # Write the cell value to the next available row in column A
        sheet.cell(row=next_row, column=1, value=cell_value)

        # Save the Excel workbook
        workbook.save(filename)

    def create_feature_function(dependent_var, expression, independent_vars):
        # Include necessary imports and comments
        
        import time
        
        layout = [
            [sg.Text("Enter Excel File Name (e.g., added_equations_list.xlsx):")],
            [sg.InputText(key='filename')],
   #         [sg.Text("Enter String to Write:")],
   #         [sg.InputText(key='cell_value')], 
            [sg.Button("Write to Excel"), sg.Button("Exit")]
        ]

        window = sg.Window("Excel Writer").Layout(layout)

        while True:
            event, values = window.Read()
            if event == sg.WIN_CLOSED or event == 'Exit':
                break
            elif event == 'Write to Excel':
                filename = values['filename']
       #         cell_value = values['cell_value']
                
                if filename:  ##  and cell_value
                    if filename.endswith(".xlsx"):
                        print("Got it")
                    else:
                        t = False
                        p_dot = 0
                        for i, p in enumerate(filename):
                            if p == '.':
                                t = True
                                p_dot = i
                        if t == True:
                            filename = filename[:p_dot]
                        
                        filename = filename + ".xlsx"
                        
                    write_to_excel(filename, dependent_var + " = " + expression)
                    sg.popup(f"Data written to {filename}", title="Success")
                else:
                    sg.popup("Both filename and cell value are required.", title="Error")

                time.sleep(3) 
                break
            
        window.close() 
        
        ind_eq = 0
        
        for indP,p in enumerate(expression):
            if p == '=':
                ind_eq = indP
                break
        
        expression = expression[(ind_eq+1):]
        
        func_body = "import sympy as sp\n"
        # func_body += "def preprocess_formula(expression):\n"
        # func_body += "    return expression.replace(\"^2\", \"**2\").replace(\"^\", \"**\")\n"
        
        # Generate the function body dynamically
        func_body += f"def {dependent_var}_feature({', '.join(independent_vars)}, expression):\n"
        func_body += "    var_dict = {var: float(val) for var, val in zip([" + ', '.join(independent_vars) + "], [" + ', '.join(independent_vars) + "])}\n"
        func_body += "    expression.replace(\"^\", \"**\")\n"
        func_body += "    processed_expression = expression\n"
     #   func_body += "    return float(sp.sympify(processed_expression, locals=var_dict))\n"
        func_body += "    return eval(processed_expression)\n"

        if f"{dependent_var}_feature" not in func_body:
            print(" -- Check extra.py file ...")
            print(f"{dependent_var}_feature")        
                        
        return func_body

    import unidecode
    import re
    
    def contains_accents_or_cedilla(name):
        # Remove accents using unidecode
        print("A")
        
        # import sys
        # sys.exit()
        
        name_without_accents = unidecode.unidecode(name)
    
        # Use regular expression to check for the presence of accents or the letter "ç"
        pattern = re.compile(r'[^\w\s]|[çÇ]', re.UNICODE)
        return bool(pattern.search(name_without_accents)) or not(name_without_accents == name)

    def add_features_gui(existing_features): 
        
        print("add_features_gui")
        
  #      import extra
        
        features = {}
        function_bodies = {}  # Store the function bodies temporarily
        
        count = 0
        
        main_loop = True

        while main_loop:
               layout = [
                   [sg.Text("Do you want to add more features ?")],
                   [sg.Checkbox("Yes", default=True, key="-YES-"), sg.Checkbox("No", key="-NO-")],
                   [sg.Button("Ok")]
               ]
        
               window = sg.Window("Add Features", layout)
        
               while True:
                   event, values = window.read()
        
                   if event == sg.WIN_CLOSED:
                       window.close()
                       return features
        
                   if event == "Ok" and (values["-YES-"] and not values["-NO-"]):
                       count += 1
                       break
        
                   if event == "Ok" and (values["-NO-"] and not values["-YES-"]):
                       main_loop = False
                       window.close()
                       print("A")
                       break
                   
               print("B")
         
      #         window.close()
                    
                    # if count > 0:
                    #     print("Here")
                    #     return features
            
               print("Main loop here: " + str(main_loop))
               
               if main_loop:

       #        if True:

                layout = [
                    [sg.Text("Feature name: "), sg.InputText(key="-FEATURE_NAME-")],
                    [sg.Text("Variables (comma-separated): "), sg.InputText(key="-VARIABLES-")],
                    [sg.Text("Variable values (comma-separated): "), sg.InputText(key="-VARIABLE_VALUES-")],
                    [sg.Text("Formula: "), sg.InputText(key="-FORMULA-"), sg.Button("Add")],
                ]
        
                window = sg.Window("Add Feature Details", layout)
                
                error_det = False
        
                while True:
                    event, values = window.read()
        
                    if event == sg.WIN_CLOSED:
                        window.close()
                        return features
        
                    feature_name = values["-FEATURE_NAME-"]
                    variables = values["-VARIABLES-"]
                    variable_values = values["-VARIABLE_VALUES-"]
                    formula = values["-FORMULA-"]
        
                    if event == "Add":
                        if not feature_name.strip():
                            sg.popup("Space for the feature name is empty!", title="Error")
                            error_det = True
                            continue
        
                        if not variables.strip():
                            sg.popup("Space for the variables is empty!", title="Error")
                            error_det = True
                            continue
        
                        if not variable_values.strip():
                            sg.popup("Space for the variable values is empty!", title="Error")
                            error_det = True
                            continue
        
                        if not formula.strip():
                            sg.popup("Blank space for the formula is empty!", title="Error")
                            error_det = True
                            continue
        
                        if feature_name in existing_features:
                            sg.popup("Feature already exists!", title="Error")
                            error_det = True
                            continue
                        
                        if contains_accents_or_cedilla(feature_name):
                            print("Here")
                            
                            # import sys
                            # sys.exit()
                            
                            sg.popup("Feature name must not contain any special characters!", title="Error")
                            error_det = True
                            continue
                        if not error_det:
                    
                            if ' ' in feature_name :
                                newFN= ""
                                for fe in feature_name:
                                    if fe != ' ':
                                      newFN += fe 
                                
                                feature_name = newFN
                                
                            # Create a dictionary mapping variable names to their values
                            var_dict = {var: float(val) for var, val in zip(variables.split(","), variable_values.split(","))}
            
                            # Extract the mathematical expression from the feature name
                            feature_parts = formula.split("=")
                            if len(feature_parts) != 2:
                                sg.popup("Invalid formula format. Use 'FeatureName = Expression'.", title="Error")
                                continue
            
                            # Evaluate the formula and check for validity
                            processed_expression = preprocess_formula(feature_parts[1].strip())
            
                            # Create a function body for the feature
                            function_body = create_feature_function(feature_name, processed_expression, variables.split(","))  ## feature_parts[0].strip()
                            function_bodies[feature_name] = function_body
                            
                            print(function_bodies)
            
                            window.close()
                            break
                        
                print("Main Loop 2")
                print(main_loop)
                window.close()
            
        print("Out")
        
        def is_file_empty(file_path):
            return os.stat(file_path).st_size == 0
        
        def clear_file_content(file_path):
            with open(file_path, 'w') as file:
                file.write("")
        
        def show_clear_dialog():
            layout = [
                [sg.Text("File extra.py is not empty.\nDo you want to clear all the code inside it?")],
                [sg.Button("Yes"), sg.Button("No")]
            ]
        
            window = sg.Window("Clear File Confirmation", layout)
        
            while True:
                event, _ = window.read()
        
                if event in (sg.WIN_CLOSED, "No"):
                    window.close()
                    return False
        
                if event == "Yes":
                    window.close()
                    return True
        
        # Check if extra.py is empty
        file_path = "extra.py"
        if not is_file_empty(file_path):
            if show_clear_dialog():
                clear_file_content(file_path)

        # Append the new function bodies to extra.py
        with open("extra.py", "a") as file:
            for function_body in function_bodies.values():
                file.write(function_body + "\n")
        
        import importlib 
        
        try:
            import extra
        except Exception as e:
            print(e) 
            
        importlib.reload(extra)


        # Manually update the globals() dictionary with the imported functions from extra.py
        for feature_name in function_bodies.keys():
            feature_function = getattr(extra, f"{feature_name}_feature")
            features[feature_name] = feature_function

        return features

    # Test the function with some existing feature names
    # existing_features = ["x", "y", "z"]
    new_features = add_features_gui(existing_features_list)

    print("New Features:")
    for feature_name, feature_function in new_features.items():
        print(f"Function name: {feature_function.__name__}, Dependent variable: {feature_name}")
    
    return new_features
